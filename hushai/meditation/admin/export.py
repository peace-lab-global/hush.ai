"""Data export utilities for admin panel (CSV/Excel)."""

from __future__ import annotations

from datetime import datetime
from typing import Any

from fastapi.responses import StreamingResponse
import io


def export_to_csv(data: list[dict[str, Any]], filename: str) -> StreamingResponse:
    """Export data to CSV format."""
    import csv

    output = io.StringIO()
    if not data:
        output.write("")
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_to_excel(
    data: list[dict[str, Any]], filename: str, sheet_name: str = "Sheet1"
) -> StreamingResponse:
    """Export data to Excel format."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "openpyxl Workbook must have an active worksheet"
    ws.title = sheet_name

    if not data:
        wb.save(io.BytesIO())
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    headers = list(data[0].keys())
    ws.append(headers)  # type: ignore[union-attr]

    for row in data:
        ws.append([_serialize_value(v) for v in row.values()])  # type: ignore[union-attr]

    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 30  # type: ignore[union-attr]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _serialize_value(value: Any) -> str:
    """Serialize a value for Excel export."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if value is None:
        return ""
    return str(value)


def get_export_response(
    data: list[dict[str, Any]],
    filename_base: str,
    export_format: str,
) -> StreamingResponse:
    """Get export response in the specified format."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if export_format == "xlsx":
        filename = f"{filename_base}_{timestamp}.xlsx"
        return export_to_excel(data, filename)
    else:
        filename = f"{filename_base}_{timestamp}.csv"
        return export_to_csv(data, filename)
